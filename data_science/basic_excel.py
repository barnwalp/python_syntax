from openpyxl import Workbook


wb = Workbook()
ws = wb.active

# create a new sheet
ws1 = wb.create_sheet("Mysheet")

print(wb.sheetnames)

# looping through worksheets
for sheet in wb:
    print(sheet.title)

ws['A1'] = 3
ws.cell(row=1, column=2, value=10)

colC = ws['C']
col_range = ws['C:D']

row10 = ws[10]
row_range = ws[5:10]

print(f'{colC} --> {col_range}')
wb.save('test.xlsx')
