from openpyxl import load_workbook
from datetime import datetime, date, time, timezone


# python datetime module
print(f'current date and time: {datetime.now()}')
print(f'Date is: {date(2005, 7, 14)}')
print(f'Time is: {time(12, 30)}')
print(f'datetime is: {datetime.combine(date(2005, 7, 14), time(12, 30))}')
print(f'datetime with timezone: {datetime.now(timezone.utc)}')
print(f'datetime: {datetime.strptime("17/06/20 08:30", "%d/%m/%y %H:%M")}')

wb = load_workbook(filename='automation mis.xlsx')

# get the worksheet to work on
ws1 = wb['Online RO list']

# Getting a value of a cell
print(f'Value of D12 cell of first worksheet is : {ws1["D12"].value}')

"""
for char in 'ABCDEFGH':
    str = char + "1"
    print(ws1[str].value)
"""
# ws4 = wb.create_sheet("testing")

# for ws in wb:
#     print(ws.title)

# ws4['A1'] = datetime.now()
# print(f'{ws4["A1"].value} --> {ws["A1"].number_format}')

# Using excel formula in python
# ws4['A2'] = "=SUM(1, 1)"

ws = wb["testing"]
# merge and unmerge in excel
ws.merge_cells('B3:D3')
ws.unmerge_cells('B3:D3')

# fold and unfold cells
# ws.column_dimensions.group('C', 'E', hidden=False)
# ws.row_dimensions.group(4, 10, hidden=False)

# Inserting Rows and columns
# This will insert a row before existing 7 rows
ws.insert_rows(7)

# Deleting columns F:I
ws.delete_cols(6, 4)

ws['C2'] = "This is a string"
wb.save('automation mis.xlsx')
