import os
from openpyxl import load_workbook
from exchangelib import Account, Message, Credentials, HTMLBody
from exchangelib import Configuration, DELEGATE

# Getting the sensitive information from environment variables
outlook_user = os.environ.get('OUTLOOK_USER')
outlook_password = os.environ.get('OUTLOOK_PASS')
outlook_server = os.environ.get('OUTLOOK_SERVER')
outlook_email = os.environ.get('OUTLOOK_EMAIL')

# Using necessary credential and config to connect exchange server
credentials = Credentials(username=outlook_user, password=outlook_password)
config = Configuration(server=outlook_server,
                       credentials=credentials)
account = Account(primary_smtp_address=outlook_email,
                  config=config, autodiscover=False,
                  access_type=DELEGATE)

path = '/mnt/c/Users/panka/Downloads/eRACTS_Complaints.xlsx'
wb = load_workbook(path)


def create_mail_table(wb):
    sheet = wb.active
    # print(sheet.title)

    # openpyxl can not copy sheet with merged cells, first all such cells
    # need to be unmerged and deleted
    # Delete first two merged rows
    sheet.unmerge_cells('A1:AI1')
    sheet.unmerge_cells('A2:AI2')
    sheet.delete_rows(1, 2)

    # Create copy mail for alteration
    ws = wb.copy_worksheet(sheet)
    ws.title = "mail"
    print(ws.title)
    ws = wb["mail"]

    # Delete column U:AH
    ws.delete_cols(21, 14)
    # Delete column P:Q
    ws.delete_cols(16, 2)
    # Delete column H:N
    ws.delete_cols(8, 7)
    # Delete column D:E
    ws.delete_cols(4, 2)

    # add a row before column 10
    ws.insert_cols(10)
    ws['J1'].value = "Detailed complaint"

    # creating a function for filter to filter out None and int types

    def is_not_str(x):
        # isinstance(x, int) will check if x is integer
        if x is None or isinstance(x, int):
            return False
        else:
            return True

    for cell in ws['J']:
        value_1 = f'G{cell.row}'
        value_2 = f'H{cell.row}'
        value_3 = f'I{cell.row}'
        merge_list = [ws[value_1].value,
                      ws[value_2].value,
                      ws[value_3].value
                      ]
        # filter will return only those value which are neither None nor int
        cell.value = ', '.join(filter(is_not_str, merge_list))

    # Deleting column G:I
    ws.delete_cols(7, 3)

    filter_value = {'Auto-Closed for Dealer Late Acknowledgement',
                    'Service Completed',
                    'Auto-Closed on Alarm Normalization'
                    }
    # Deleting all complaints which have been closed
    for cell in ws['H']:
        if cell.row > 1:
            if cell.value in filter_value:
                print(cell.value)
                ws.delete_rows(cell.row)
    # Delete end column H
    ws.delete_cols(8)
    ws['H1'] = "Current Status of complaint"
    ws['I1'] = "Target Date"

    wb.save(path)
    return wb


def create_html(wb):
    ws = wb["mail"]

    # contains content in HTML format
    html = """
    <html><body><h3>Pl find the current list of eRACTS complaint which have not yet been resolved, pl go through each case and provide current status of complaint and target date of rectification:</h3>
    """

    for whole_row in ws.iter_rows(min_row=1, max_row=1):
        html += "<table border=1><tr>"
        for row in whole_row:
            html += f'<th>{row.value}</th>'
        html += "</tr>"

    for whole_row in ws.iter_rows(min_row=2):
        html += "<tr>"
        for row in whole_row:
            if row.value is not None:
                html += f'<td>{row.value}</td>'
            else:
                html += f'<td>{" "}</td>'
        html += "</tr>"

    html += """
    </table>
    <p>
    <br>
    With Regards<br>
    Pankaj Barnwal<br>
    </p>
    </body>
    </html>
    """
    # print(html)
    return html


wb = create_mail_table(wb)
msg = Message(
    account=account,
    subject="eRACTS complaint status - Sambalpur DO",
    body=HTMLBody(create_html(wb)),
    to_recipients=['barnwalp@indianoil.in']
)

msg.send_and_save()

"""
with open("Simple.html", "w", encoding="utf-8") as filehtml:
    filehtml.write(html)

os.system("Simple.html")
"""
